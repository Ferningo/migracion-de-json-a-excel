from openpyxl import Workbook
from json_loader import obtener_alias
from json_loader import obtener_queries

wb = Workbook()

ws = wb.active

ws["A1"] = "ALIAS"
ws["B1"] = "QUERY"


index = 2

aliases = obtener_alias('9427.json')
queries = obtener_queries('9427.json')


for alias in aliases:
    ws[f"A{index}"] = alias
    ws[f"B{index}"] = queries[index-2]
    index += 1


# Save the file
wb.save("excel_final.xlsx")
